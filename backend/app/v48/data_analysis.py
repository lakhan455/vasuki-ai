from __future__ import annotations

import csv
import io
import json
import math
import statistics
from collections import Counter
from typing import Any

MAX_FILE_BYTES = 10 * 1024 * 1024
MAX_ROWS = 5000
MAX_COLUMNS = 100
PREVIEW_ROWS = 20


def _clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value).strip()
    return text if text else None


def _as_number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        try:
            return float(text[:-1]) / 100.0
        except ValueError:
            return None
    try:
        number = float(text)
        return number if math.isfinite(number) else None
    except ValueError:
        return None


def _normalise_rows(headers: list[str], rows: list[list[Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    safe_headers: list[str] = []
    seen: Counter[str] = Counter()
    for index, header in enumerate(headers[:MAX_COLUMNS]):
        base = str(header or "").strip() or f"column_{index + 1}"
        seen[base] += 1
        safe_headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")

    records: list[dict[str, Any]] = []
    for raw in rows[:MAX_ROWS]:
        padded = list(raw[: len(safe_headers)]) + [None] * max(0, len(safe_headers) - len(raw))
        records.append({key: _clean_cell(value) for key, value in zip(safe_headers, padded)})
    return safe_headers, records


def _parse_csv(data: bytes) -> tuple[list[str], list[dict[str, Any]], str]:
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = list(reader)
    if not rows:
        return [], [], "csv"
    headers, records = _normalise_rows(rows[0], rows[1:])
    return headers, records, "csv"


def _parse_json(data: bytes) -> tuple[list[str], list[dict[str, Any]], str]:
    parsed = json.loads(data.decode("utf-8-sig", errors="strict"))
    if isinstance(parsed, dict):
        for key in ("rows", "data", "items", "records"):
            if isinstance(parsed.get(key), list):
                parsed = parsed[key]
                break
        else:
            parsed = [parsed]
    if not isinstance(parsed, list):
        raise ValueError("JSON must contain an object or an array of records.")

    records_raw = [item for item in parsed[:MAX_ROWS] if isinstance(item, dict)]
    headers: list[str] = []
    for item in records_raw:
        for key in item.keys():
            value = str(key)
            if value not in headers:
                headers.append(value)
            if len(headers) >= MAX_COLUMNS:
                break
        if len(headers) >= MAX_COLUMNS:
            break
    records = [{key: _clean_cell(item.get(key)) for key in headers} for item in records_raw]
    return headers, records, "json"


def _parse_xlsx(data: bytes) -> tuple[list[str], list[dict[str, Any]], str]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - dependency contract
        raise RuntimeError("XLSX analysis requires openpyxl.") from exc

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = list(next(iterator))
    except StopIteration:
        return [], [], "xlsx"
    rows: list[list[Any]] = []
    for index, row in enumerate(iterator):
        if index >= MAX_ROWS:
            break
        rows.append(list(row))
    headers, records = _normalise_rows(header_row, rows)
    return headers, records, "xlsx"


def _profile_column(name: str, records: list[dict[str, Any]]) -> dict[str, Any]:
    values = [row.get(name) for row in records]
    non_missing = [value for value in values if value is not None and str(value).strip() != ""]
    numeric = [number for value in non_missing if (number := _as_number(value)) is not None]
    unique = len({json.dumps(value, ensure_ascii=False, sort_keys=True, default=str) for value in non_missing})
    profile: dict[str, Any] = {
        "name": name,
        "non_missing": len(non_missing),
        "missing": len(values) - len(non_missing),
        "unique": unique,
        "numeric_ratio": round(len(numeric) / max(1, len(non_missing)), 4),
    }
    if non_missing and len(numeric) >= max(2, int(len(non_missing) * 0.8)):
        profile.update(
            {
                "type": "numeric",
                "min": round(min(numeric), 6),
                "max": round(max(numeric), 6),
                "mean": round(statistics.fmean(numeric), 6),
                "median": round(statistics.median(numeric), 6),
            }
        )
    else:
        counts = Counter(str(value) for value in non_missing)
        profile.update(
            {
                "type": "categorical",
                "top_values": [
                    {"value": value, "count": count}
                    for value, count in counts.most_common(5)
                ],
            }
        )
    return profile


def _chart_suggestions(columns: list[dict[str, Any]]) -> list[dict[str, str]]:
    numeric = [column["name"] for column in columns if column.get("type") == "numeric"]
    categorical = [column["name"] for column in columns if column.get("type") == "categorical"]
    suggestions: list[dict[str, str]] = []
    if categorical and numeric:
        suggestions.append({"type": "bar", "x": categorical[0], "y": numeric[0]})
    if len(numeric) >= 2:
        suggestions.append({"type": "scatter", "x": numeric[0], "y": numeric[1]})
    if numeric:
        suggestions.append({"type": "histogram", "x": numeric[0], "y": "count"})
    return suggestions[:3]


def analyze_tabular_bytes(filename: str, data: bytes) -> dict[str, Any]:
    if len(data) > MAX_FILE_BYTES:
        raise ValueError("Data file exceeds the 10 MB V48 analysis limit.")
    lower = (filename or "").lower()
    if lower.endswith(".csv") or lower.endswith(".tsv"):
        headers, records, kind = _parse_csv(data)
    elif lower.endswith(".json"):
        headers, records, kind = _parse_json(data)
    elif lower.endswith(".xlsx"):
        headers, records, kind = _parse_xlsx(data)
    else:
        raise ValueError("Supported data files: .csv, .tsv, .json and .xlsx")

    columns = [_profile_column(name, records) for name in headers]
    missing_cells = sum(int(column.get("missing") or 0) for column in columns)
    return {
        "ok": True,
        "version": "v48",
        "file": filename,
        "format": kind,
        "rows": len(records),
        "columns_count": len(headers),
        "missing_cells": missing_cells,
        "columns": columns,
        "preview": records[:PREVIEW_ROWS],
        "chart_suggestions": _chart_suggestions(columns),
        "limits": {
            "max_bytes": MAX_FILE_BYTES,
            "max_rows_profiled": MAX_ROWS,
            "max_columns": MAX_COLUMNS,
        },
    }


def spreadsheet_text(filename: str, data: bytes, max_chars: int = 80000) -> str:
    report = analyze_tabular_bytes(filename, data)
    compact = {
        "file": report["file"],
        "format": report["format"],
        "rows": report["rows"],
        "columns": report["columns"],
        "preview": report["preview"],
    }
    return json.dumps(compact, ensure_ascii=False, default=str)[:max_chars]
